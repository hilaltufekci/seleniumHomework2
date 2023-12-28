from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
from selenium.webdriver.support.wait import WebDriverWait 
from selenium.webdriver.support import expected_conditions as ec 
import pytest
import openpyxl
from constants import globalConstants as c

class Test_CaseHomework2:

    def setup_method(self):#her test başlangıcında çalışıcak fonk.
        self.driver=webdriver.Chrome()
        self.driver.get(c.BASE_URL)
        self.driver.maximize_window()

    def teardown_method(self): #her testin bitiminde çalışıcak fonk
        self.driver.quit()
    
    def getData():
        excel = openpyxl.load_workbook(c.invalid_login_xlsx)
        sheet = excel["Sayfa1"] #hangi sayfada çalışacağımı gösteriyorum
        rows = sheet.max_row #kaçıncı satıra kadar veri var?
        data = []
        for i in range(2,rows+1):
            username = sheet.cell(i,1).value
            password = sheet.cell(i,2).value
            data.append((username,password))

        return data
    
    def getData2():
        excel = openpyxl.load_workbook(c.name_code_list_xlsx)
        sheet = excel["Sayfa1"] #hangi sayfada çalışacağımı gösteriyorum
        rows = sheet.max_row #kaçıncı satıra kadar veri var?
        data = []
        for i in range(2,rows+1):
            firstname = sheet.cell(i,1).value
            lastname = sheet.cell(i,2).value
            postalcode=sheet.cell(i,3).value
            data.append((firstname,lastname,postalcode))

        return data


    def test_invalid_login1(self):
        usernameInput=self.driver.find_element(By.ID,c.USERNAME_ID).send_keys("locked_out_user")
        passwordInput=self.driver.find_element(By.ID,c.PASSWORD_ID).send_keys("1234")
        loginButton=self.driver.find_element(By.ID,c.LOGIN_BUTTON_ID).click()
        errorMessage=self.driver.find_element(By.XPATH,"//*[@id='login_button_container']/div/form/div[3]/h3")
        assert errorMessage.text=="Epic sadface: Username and password do not match any user in this service"

    @pytest.mark.parametrize("username,password",getData())
    def test_invalid_login2(self,username,password):
        usernameInput=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.USERNAME_ID))).send_keys(username)
        passwordInput=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.PASSWORD_ID))).send_keys(password)
        loginButton=self.driver.find_element(By.ID,c.LOGIN_BUTTON_ID).click()
        headerLogo=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.HEADER_LOGO_XPATH)))
        assert headerLogo.text =="Swag Labs" 
        sleep(5)

  
    @pytest.mark.parametrize("firstname,lastname,postalcode",getData2())
    def test_cart_control(self,firstname,lastname,postalcode):
        usernameInput=self.driver.find_element(By.ID,c.USERNAME_ID).send_keys("standard_user")
        passwordInput=self.driver.find_element(By.ID,c.PASSWORD_ID).send_keys("secret_sauce")
        loginButton=self.driver.find_element(By.ID,c.LOGIN_BUTTON_ID).click()
        headerLogo=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.HEADER_LOGO_XPATH)))
        assert headerLogo.text =="Swag Labs"     
        self.driver.execute_script("window.scrollTo(0,500)")
        addToCart=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,"//*[@id='add-to-cart-test.allthethings()-t-shirt-(red)']"))).click()
        shoppingCart=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,"//*[@id='shopping_cart_container']/a"))).click()
        checkOutButton=self.driver.find_element(By.ID,c.CHECKOUT_BUTTON_ID).click()
        firstnameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.FIRSTNAME_ID))).send_keys(firstname)
        lastnameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.LASTNAME_ID))).send_keys(lastname)
        sleep(5)
        postalcodeInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.POSTALCODE_ID))).send_keys(postalcode)
        continueButton=self.driver.find_element(By.ID,c.CONTİNUE_BUTTON_ID).click()
        headerCheck=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH," //*[@id='header_container']/div[2]/span ")))
        assert headerCheck.text =="Checkout: Overview"
        finishButton = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.FINISH_BUTTON_ID))).click() 
        sleep (5)

   